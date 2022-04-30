import openpyxl
import os

def getTestData(id):
    data = {}
    path = os.path.relpath(os.getcwd())

    workbook = openpyxl.load_workbook(path+"\\..\\TestData\\testdata.xlsx")
    sheet = workbook.active
    for i in range(1, sheet.max_row + 1):
        if sheet.cell(row=i, column=1).value == id:
            for j in range(2, sheet.max_column + 1):
                data[sheet.cell(row=1,column=j).value] = sheet.cell(row=i, column=j).value
    return [data]